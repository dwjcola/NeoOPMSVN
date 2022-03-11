#coding=utf-8
# from __future__ import absolute_import
# Copyright (c) 2010-2017 openpyxl
"""
lixu 20171109
Monkey patching for openpyxl.load_workbook
Without parsing style.xml in xlsx files, loading process can be much much faster.
"""
from openpyxl.reader.workbook import WorkbookParser
from io import BytesIO
from warnings import warn
import re

# compatibility imports
from openpyxl.xml.functions import iterparse

# package imports
from openpyxl.cell import Cell
from openpyxl.worksheet.filters import AutoFilter, SortState
from openpyxl.cell.text import Text
from openpyxl.worksheet.dimensions import (
    ColumnDimension,
    RowDimension,
    SheetFormatProperties,
)
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.merge import MergeCells
from openpyxl.worksheet.page import PageMargins, PrintOptions, PrintPageSetup
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.views import SheetViewList
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.xml.constants import (
    SHEET_MAIN_NS,
    REL_NS,
    EXT_TYPES,
    PKG_REL_NS
)
from openpyxl.styles import Color
from openpyxl.formatting import Rule
from openpyxl.formatting.formatting import ConditionalFormatting
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.properties import WorksheetProperties
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string,
    coordinate_to_tuple,
    )
from openpyxl.descriptors.excel import ExtensionList, Extension
from openpyxl.worksheet.table import TablePartList
from openpyxl.worksheet.pagebreak import PageBreak

FLOAT_REGEX = re.compile(r"\.|[E-e]")


def _cast_number(value):
    "Convert numbers as string to an int or float"
    m = FLOAT_REGEX.search(value)
    if m is not None:
        return float(value)
    return int(value)

def safe_iterator(node, tag=None):
    """Return an iterator that is compatible with Python 2.6"""
    if node is None:
        return []
    if hasattr(node, "iter"):
        return node.iter(tag)
    else:
        return node.getiterator(tag)


def _get_xml_iter(xml_source):
    """
    Possible inputs: strings, bytes, members of zipfile, temporary file
    Always return a file like object
    """
    if not hasattr(xml_source, 'read'):
        try:
            xml_source = xml_source.encode("utf-8")
        except (AttributeError, UnicodeDecodeError):
            pass
        return BytesIO(xml_source)
    else:
        try:
            xml_source.seek(0)
        except:
            # could be a zipfile
            pass
        return xml_source

class WorkSheetParserWOS(WorkbookParser):
    """docstring for WorkbookParserWOS"""
    def __init__(self, ws, xml_source, shared_strings):
        WorkbookParser.__init__(self, ws, xml_source, shared_strings)

    def parse(self):
        dispatcher = {
            '{%s}mergeCells' % SHEET_MAIN_NS: self.parse_merge,
            '{%s}col' % SHEET_MAIN_NS: self.parse_column_dimensions,
            '{%s}row' % SHEET_MAIN_NS: self.parse_row,
            # '{%s}conditionalFormatting' % SHEET_MAIN_NS: self.parser_conditional_formatting,
            '{%s}legacyDrawing' % SHEET_MAIN_NS: self.parse_legacy_drawing,
            '{%s}sheetProtection' % SHEET_MAIN_NS: self.parse_sheet_protection,
            '{%s}extLst' % SHEET_MAIN_NS: self.parse_extensions,
            '{%s}hyperlink' % SHEET_MAIN_NS: self.parse_hyperlinks,
            '{%s}tableParts' % SHEET_MAIN_NS: self.parse_tables,
                      }

        properties = {
            '{%s}printOptions' % SHEET_MAIN_NS: ('print_options', PrintOptions),
            '{%s}pageMargins' % SHEET_MAIN_NS: ('page_margins', PageMargins),
            '{%s}pageSetup' % SHEET_MAIN_NS: ('page_setup', PrintPageSetup),
            '{%s}headerFooter' % SHEET_MAIN_NS: ('HeaderFooter', HeaderFooter),
            '{%s}autoFilter' % SHEET_MAIN_NS: ('auto_filter', AutoFilter),
            '{%s}dataValidations' % SHEET_MAIN_NS: ('data_validations', DataValidationList),
            #'{%s}sheet/{%s}sortState' % (SHEET_MAIN_NS, SHEET_MAIN_NS): ('sort_state', SortState),
            '{%s}sheetPr' % SHEET_MAIN_NS: ('sheet_properties', WorksheetProperties),
            '{%s}sheetViews' % SHEET_MAIN_NS: ('views', SheetViewList),
            '{%s}sheetFormatPr' % SHEET_MAIN_NS: ('sheet_format', SheetFormatProperties),
            '{%s}rowBreaks' % SHEET_MAIN_NS: ('page_breaks', PageBreak),
        }

        tags = dispatcher.keys()
        stream = _get_xml_iter(self.source)
        it = iterparse(stream, tag=tags)

        for _, element in it:
            tag_name = element.tag
            if tag_name in dispatcher:
                dispatcher[tag_name](element)
                element.clear()
            elif tag_name in properties:
                prop = properties[tag_name]
                obj = prop[1].from_tree(element)
                setattr(self.ws, prop[0], obj)
                element.clear()

        self.ws._current_row = self.ws.max_row

    def parse_cell(self, element):
        value = element.find(self.VALUE_TAG)
        if value is not None:
            value = value.text
        formula = element.find(self.FORMULA_TAG)
        data_type = element.get('t', 'n')
        coordinate = element.get('r')
        self._col_count += 1
        # lixu 20171109
        # style_id = element.get('s')
        style_id=None
        # assign formula to cell value unless only the data is desired
        if formula is not None and not self.data_only:
            data_type = 'f'
            if formula.text:
                value = "=" + formula.text
            else:
                value = "="
            formula_type = formula.get('t')
            if formula_type:
                if formula_type != "shared":
                    self.ws.formula_attributes[coordinate] = dict(formula.attrib)

                else:
                    si = formula.get('si')  # Shared group index for shared formulas
                    if si in self.shared_formula_masters:
                        trans = self.shared_formula_masters[si]
                        value = trans.translate_formula(coordinate)
                    else:
                        self.shared_formula_masters[si] = Translator(value, coordinate)

        style_array = None
        if style_id is not None:
            style_id = int(style_id)
            style_array = self.styles[style_id]

        if coordinate:
            row, column = coordinate_to_tuple(coordinate)
        else:
            row, column = self._row_count, self._col_count

        cell = Cell(self.ws, row=row, col_idx=column, style_array=style_array)
        self.ws._cells[(row, column)] = cell

        if value is not None:
            if data_type == 'n':
                value = _cast_number(value)
            elif data_type == 'b':
                value = bool(int(value))
            elif data_type == 's':
                value = self.shared_strings[int(value)]
            elif data_type == 'str':
                data_type = 's'

        else:
            if data_type == 'inlineStr':
                child = element.find(self.INLINE_STRING)
                if child is not None:
                    data_type = 's'
                    richtext = Text.from_tree(child)
                    value = richtext.content

        if self.guess_types or value is None:
            cell.value = value
        else:
            cell._value=value
            cell.data_type=data_type

    def parse_column_dimensions(self, col):
        attrs = dict(col.attrib)
        column = get_column_letter(int(attrs['min']))
        attrs['index'] = column
        # lixu 20171109
        # if 'style' in attrs:
        #     attrs['style'] = self.styles[int(attrs['style'])]
        attrs['style']=None
        dim = ColumnDimension(self.ws, **attrs)
        self.ws.column_dimensions[column] = dim

    def parse_row(self, row):
        attrs = dict(row.attrib)

        if "r" in attrs:
            self._row_count = int(attrs['r'])
        else:
            self._row_count += 1
        self._col_count = 0
        keys = set(attrs)
        for key in keys:
            if key == "s":
                # lixu 20171109
                # attrs['s'] = self.styles[int(attrs['s'])]
                attrs['s']=None
            elif key.startswith('{'):
                del attrs[key]


        keys = set(attrs)
        if keys != set(['r', 'spans']) and keys != set(['r']):
            # don't create dimension objects unless they have relevant information
            dim = RowDimension(self.ws, **attrs)
            self.ws.row_dimensions[dim.index] = dim

        for cell in safe_iterator(row, self.CELL_TAG):
            self.parse_cell(cell)
            
    def parser_conditional_formatting(self, element):
        cf = ConditionalFormatting.from_tree(element)
        for rule in cf.rules:
            rule.dxfId = None
            self.ws.conditional_formatting.add(cf.sqref, rule)
		
		





"""Read an xlsx file into Python"""

# Python stdlib imports
from zipfile import ZipFile, ZIP_DEFLATED, BadZipfile
from sys import exc_info
from io import BytesIO
import os.path
import warnings

# compatibility imports

# Allow blanket setting of KEEP_VBA for testing
KEEP_VBA = False
# try:
#     from ..tests import KEEP_VBA
# except ImportError:
#     KEEP_VBA = False


# package imports
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.xml.constants import (
    ARC_SHARED_STRINGS,
    ARC_CORE,
    ARC_CONTENT_TYPES,
    ARC_WORKBOOK,
    ARC_THEME,
    COMMENTS_NS,
    SHARED_STRINGS,
    EXTERNAL_LINK,
    XLTM,
    XLTX,
    XLSM,
    XLSX,
)

from openpyxl.comments.comment_sheet import CommentSheet
from openpyxl.workbook import Workbook

from openpyxl.reader.strings import read_string_table
from openpyxl.styles.stylesheet import apply_stylesheet

from openpyxl.packaging.core import DocumentProperties
from openpyxl.packaging.manifest import Manifest
from openpyxl.reader.workbook import WorkbookParser
from openpyxl.packaging.relationship import get_dependents, get_rels_path

from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.table import Table

from openpyxl.xml.functions import fromstring

# Use exc_info for Python 2 compatibility with "except Exception[,/ as] e"


CENTRAL_DIRECTORY_SIGNATURE = b'\x50\x4b\x05\x06'
SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm')


def repair_central_directory(zipFile, is_file_instance):
    ''' trims trailing data from the central directory
    code taken from http://stackoverflow.com/a/7457686/570216, courtesy of Uri Cohen
    '''

    f = zipFile if is_file_instance else open(zipFile, 'rb+')
    data = f.read()
    pos = data.find(CENTRAL_DIRECTORY_SIGNATURE)  # End of central directory signature
    if (pos > 0):
        sio = BytesIO(data)
        sio.seek(pos + 22)  # size of 'ZIP end of central directory record'
        sio.truncate()
        sio.seek(0)
        return sio

    f.seek(0)
    return f



def _validate_archive(filename):
    """
    Check the file is a valid zipfile
    """
    is_file_like = hasattr(filename, 'read')

    if not is_file_like and os.path.isfile(filename):
        file_format = os.path.splitext(filename)[-1].lower()
        if file_format not in SUPPORTED_FORMATS:
            if file_format == '.xls':
                msg = ('openpyxl does not support the old .xls file format, '
                       'please use xlrd to read this file, or convert it to '
                       'the more recent .xlsx file format.')
            elif file_format == '.xlsb':
                msg = ('openpyxl does not support binary format .xlsb, '
                       'please convert this file to .xlsx format if you want '
                       'to open it with openpyxl')
            else:
                msg = ('openpyxl does not support %s file format, '
                       'please check you can open '
                       'it with Excel first. '
                       'Supported formats are: %s') % (file_format,
                                                       ','.join(SUPPORTED_FORMATS))
            raise InvalidFileException(msg)


    if is_file_like:
        # fileobject must have been opened with 'rb' flag
        # it is required by zipfile
        if getattr(filename, 'encoding', None) is not None:
            raise IOError("File-object must be opened in binary mode")

    try:
        archive = ZipFile(filename, 'r', ZIP_DEFLATED)
    except BadZipfile:
        f = repair_central_directory(filename, is_file_like)
        archive = ZipFile(f, 'r', ZIP_DEFLATED)
    return archive


def _find_workbook_part(package):
    for ct in [XLTM, XLTX, XLSM, XLSX]:
        part = package.find(ct)
        if part:
            return part

    raise IOError("File contains no valid workbook part")


def load_workbook(filename, read_only=False, keep_vba=KEEP_VBA,
                  data_only=False, guess_types=False, keep_links=True):
    """Open the given filename and return the workbook

    :param filename: the path to open or a file-like object
    :type filename: string or a file-like object open in binary mode c.f., :class:`zipfile.ZipFile`

    :param read_only: optimised for reading, content cannot be edited
    :type read_only: bool

    :param keep_vba: preseve vba content (this does NOT mean you can use it)
    :type keep_vba: bool

    :param guess_types: guess cell content type and do not read it from the file
    :type guess_types: bool

    :param data_only: controls whether cells with formulae have either the formula (default) or the value stored the last time Excel read the sheet
    :type data_only: bool

    :param keep_links: whether links to external workbooks should be preserved. The default is True
    :type keep_links: bool

    :rtype: :class:`openpyxl.workbook.Workbook`

    .. note::

        When using lazy load, all worksheets will be :class:`openpyxl.worksheet.iter_worksheet.IterableWorksheet`
        and the returned workbook will be read-only.

    """
    from datetime import datetime
    # print "running Monkey patch"
    archive = _validate_archive(filename)
    read_only = read_only

    src = archive.read(ARC_CONTENT_TYPES)
    root = fromstring(src)
    package = Manifest.from_tree(root)

    wb_part = _find_workbook_part(package)
    parser = WorkbookParser(archive, wb_part.PartName[1:])
    wb = parser.wb
    wb._data_only = data_only
    wb._read_only = read_only
    wb._keep_links = keep_links
    wb.guess_types = guess_types
    wb.template = wb_part.ContentType in (XLTX, XLTM)
    parser.parse()
    wb._sheets = []

    if read_only and guess_types:
        warnings.warn('Data types are not guessed when using iterator reader')

    valid_files = archive.namelist()
    
    
    # If are going to preserve the vba then attach a copy of the archive to the
    # workbook so that is available for the save.
    if keep_vba:
        wb.vba_archive = ZipFile(BytesIO(), 'a', ZIP_DEFLATED)
        for name in archive.namelist():
            wb.vba_archive.writestr(name, archive.read(name))


    if read_only:
        wb._archive = ZipFile(filename)

    # get workbook-level information
    if ARC_CORE in valid_files:
        src = fromstring(archive.read(ARC_CORE))
        wb.properties = DocumentProperties.from_tree(src)


    shared_strings = []
    ct = package.find(SHARED_STRINGS)
    if ct is not None:
        strings_path = ct.PartName[1:]
        shared_strings = read_string_table(archive.read(strings_path))


    if ARC_THEME in valid_files:
        wb.loaded_theme = archive.read(ARC_THEME)
    # lixu
    start = datetime.now()
    # apply_stylesheet(archive, wb) # bind styles to workbook
    # print "openpyxl apply_stylesheet in  load_workbook"+str(datetime.now()-start)
    # get worksheets
    for sheet, rel in parser.find_sheets():
        sheet_name = sheet.name
        worksheet_path = rel.target
        rels_path = get_rels_path(worksheet_path)
        rels = []
        if rels_path in valid_files:
            rels = get_dependents(archive, rels_path)

        if not worksheet_path in valid_files:
            continue

        if read_only:
            ws = ReadOnlyWorksheet(wb, sheet_name, worksheet_path, None,
                                   shared_strings)

            wb._sheets.append(ws)
        else:
            fh = archive.open(worksheet_path)
            ws = wb.create_sheet(sheet_name)
            ws._rels = rels
            ws_parser = WorkSheetParserWOS(ws, fh, shared_strings)
            ws_parser.parse()

            if rels:
                # assign any comments to cells
                for r in rels.find(COMMENTS_NS):
                    src = archive.read(r.target)
                    comment_sheet = CommentSheet.from_tree(fromstring(src))
                    for ref, comment in comment_sheet.comments:
                        ws[ref].comment = comment

                # preserve link to VML file if VBA
                if (
                    wb.vba_archive is not None
                    and ws.legacy_drawing is not None
                    ):
                    ws.legacy_drawing = rels[ws.legacy_drawing].target

                for t in ws_parser.tables:
                    src = archive.read(t)
                    xml = fromstring(src)
                    table = Table.from_tree(xml)
                    ws.add_table(table)

        ws.sheet_state = sheet.state
        ws._rels = [] # reset

    parser.assign_names()

    #wb._differential_styles.styles =  [] # tables may depened upon dxf

    archive.close()
    return wb

import openpyxl.reader.excel as oe
oe.load_workbook = load_workbook
