#!/usr/bin/python3
#coding=utf-8
#author: cody
"""
libiqi 20220111
由于梦伽配表格式和表头与topjoy不同做了些修改，且由于openpyxl版本更新不再使用之前的LoadWorkbookWithoutStyles
"""
import time
import json
import re
import openpyxl as lwws
from Core.Container import Container
from Core.ExcelRecord import ExcelRecord

LOCATION_ROW  = 5
TYPE_DESC_ROW = 4
KEY_ROW       = 5
TYPE_ROW      = 3
OPTION_ROW    = 6
LINK_ROW      = 1
VALUE_ROW     = 8
KEY_COLUMN    = 1
IGNORE_SHEET_PREFIX = "$"
MAX_COLUMN = 100

def reshape_list (exlist,a):
    """
    一行list变成a列list
    """
    if len(exlist)//a <=1:
        return exlist
    result = []
    for n in range(len(exlist)//a):
        for i in range(a):
            if i == 0:
                result.append([])
            result[n].append(exlist[i+n*a])
    return result

class DreamerExcel:
    def __init__(self, table_name, excel_name):
        self._work_book = lwws.load_workbook(filename=excel_name, read_only=False, data_only=True)
        self._schema = {}
        self._link_info = []
        # 过滤掉以$开头的sheet
        self._valid_sheet = []
        self._model = {}                                                                                                                      
        self._table_name = table_name
        # 用于加载数据异常时定位错误数据并产出报告
        self._current_sheet = ""
        self._current_key = ""
        self._current_tag = ""
        self._exception_info = []
        self._chain = [] # Traslate 表被拆分了很多的表，但是link的名称都是Translate
        self._all_tags = []
        self._all_records = []
        self._value_dict= {}
        self._preload()

    def _preload(self):
        for sheet_name in self._work_book.get_sheet_names():
            if sheet_name.find(IGNORE_SHEET_PREFIX) >= 0:
                continue
            self._valid_sheet.append(sheet_name)
        sheet_name = self._valid_sheet[0]
        self._init_schema()

    def _get_valid_column(self, sheet_name):
        """ 获取有效的location行，用于确认那些列是有效列"""
        location_cells = self.query_cells_by_row(sheet_name, LOCATION_ROW)
        return [cell.col_idx for cell in location_cells if cell.value != None and cell.value != "$"]

    def _init_schema(self):
        """ 每个sheet的列排列可能有差异，需要存储key和type的关系"""
        link_dict = {}
        for sheet_name in self._valid_sheet:
            self._current_sheet = sheet_name
            valid_cols = self._get_valid_column(sheet_name)
            work_sheet = self._work_book.get_sheet_by_name(sheet_name)
            sheet_schema = {}
            key_list = []
            option_cell = work_sheet.cell(row=OPTION_ROW, column=KEY_COLUMN)
            # if option_cell.value != None and option_cell.value.find("__option") >=0:
            #     self._value_dict[sheet_name] = OPTION_ROW + 2
            # else:
            self._value_dict[sheet_name] = OPTION_ROW

            for col_idx in valid_cols:
                key_cell = work_sheet.cell(row=KEY_ROW, column=col_idx)
                key_val = key_cell.value
                self._current_key = key_val
                if key_val == None: # 策划飘逸的配置
                    self.raise_exception("Key None column: {} is None".format(col_idx))
                    continue
                # xx(a,b) -> ['xx','a,b','']
                key_rule = r"[\(\)]"
                re_key_group = re.split(key_rule, key_val)
                key_schema = {
                    "column":key_cell.col_idx,
                }
                if len(re_key_group) > 1:
                    # 包含「()」是个复杂字段例如 position(x,y)
                    key_schema["special_key"] =re_key_group[1].split(",")

                link_cell = work_sheet.cell(row=LINK_ROW, column=col_idx)
                link_val = link_cell.value
                if link_val != None and str(link_val).find("link:") >= 0:
                    link_val = link_val[5:]
                    link_dict[key_val] = link_val

                type_cell = work_sheet.cell(row=TYPE_ROW, column=col_idx)
                type_rule = r"[\[\]]"
                type_str = type_cell.value
                if type_str == None:
                    self.raise_exception("Type None column:{} is None".format(col_idx))
                    type_str = "string"
                    
                # 复杂结构 [string], [string,int,float]
                re_type_group = re.split(type_rule, type_str)
                if len(re_type_group) > 1:
                    key_schema["type"] = re_type_group[1].split(",")
                    key_schema["need_remove_bra"] = 1
                else:
                    key_schema["type"] = re_type_group[0]
                sheet_schema[re_key_group[0]] = key_schema
                key_list.append(re_key_group[0])
                sheet_schema["key_list"] = key_list
            self._init_record(sheet_name, sheet_schema)
            self._schema[sheet_name] = sheet_schema
        self._schema["link_table"] = link_dict

    def _init_record(self, sheet_name, sheet_schema):
        work_sheet = self._work_book.get_sheet_by_name(sheet_name)
        self._current_sheet = sheet_name
        value_row = self._value_dict.get(sheet_name)
        # is_start = False
        is_end = False
        for i in range(value_row, work_sheet.max_row + 1):
            if is_end:
                break
            # flag_cell = work_sheet.cell(row=i, column=1)
            # if flag_cell.value == "start":
            #     is_start = True
            # elif flag_cell.value == "end":
            #     is_end = True
            # if not is_start:
            #     continue
            is_empty = False
            data_dict = {}
            for key in sheet_schema.get("key_list"):
                self._current_key = key
                key_schema = sheet_schema.get(key)
                cell = work_sheet.cell(row=i, column=key_schema.get("column"))
                if key == "id":
                    self._current_tag = cell.value
                    if cell.value == None:
                        # 空行数据跳过
                        is_empty = True
                        is_end = True
                        # 梦伽HA项目需求有一行id为空后面的就都不导表了，相当于之前tkw的配置了end
                        break
                try:
                    value = self.format_cell(key_schema, cell.value)
                except Exception as e:
                    value = None
                    self.raise_exception("Parse Error {}".format(str(e)))
                data_dict[key] = value
            if not is_empty:
                record = ExcelRecord(self, data_dict)
                self._model[record.tag] = record
                self._all_tags.append(record.tag)
                self._all_records.append(record)

    def query_cells_by_row(self, sheet_name, row_num):
        """ 根据sheet名字和行号获取整行数据 """
        limit = {
            "min_row":row_num,
            "max_row":row_num
        }

        work_sheet = self._work_book.get_sheet_by_name(sheet_name)
        # 这里最大列只有50
        # 策划的一些非常规操作会导致表里有很多空列
        if work_sheet.max_column > MAX_COLUMN:
            limit["max_col"] = MAX_COLUMN
            print(work_sheet.max_column, "max colomn is out of limit", self._table_name,sheet_name)
        # TODO 这种错误的表需要过滤出来修正

        for row in work_sheet.iter_rows(**limit):
            # 这里只会有一行数据
            return row

    def format_string(self, value):
        return str(value)

    def format_int(self, value):
        value = value.strip()
        if value == "":
            return
        return int(value)
    
    def format_float(self, value):
        return float(value)

    def format_double(self, value):
        return float(value)

    def format_json(self, value):
        return json.loads(value)

    def format_array(self, value):
        return json.loads(value)

    def format_dispatch(self, type_str, value_str):
        func = getattr(self, "format_" + type_str)
        return func(value_str)

    def format_complex_data(self, key_list, type_list, value_str):
        # ext_key  ["name", "age"]
        # type_str [string,int]
        # value_str cody,18;coco,3
        # TODO 有的数组会在最后多加一个";"
        value_ret = []
        # key_rule = r"[\[\]]"
        # re_key_group = re.split(key_rule, value_str)
        # i = 1
        # value_str_list = []
        # while i<len(re_key_group):
        #     value_str_list.append(re_key_group[i])
        #     i+=2
        for values in value_str.split(";"):
            value_dict = {}
            if values == "":
                continue
            value_list = values.split(",")
            for i in range(len(type_list)):
                tmp_dict = self.format_simple_data(key_list[i], type_list[i], value_list[i])
                value_dict.update(tmp_dict)
            value_ret.append(value_dict)
        return value_ret

    # int
    # float
    # json
    # string
    # array
    def format_simple_data(self, key_str, type_str, value_str):
        return {key_str:self.format_dispatch(type_str, value_str)}

    def format_cell(self, schema, value):
        # print(key_str, type_str, value_str)
        if value == None:
            return 
        value_str = str(value)
        if schema.get("need_remove_bra") != None:
            # 梦伽配表格式和topjoy不同，用[]代替；需要去除
            key_rule = r"[\[\]]"
            re_key_group = re.split(key_rule, value_str)
            i = 1
            value_str = ""
            while i<len(re_key_group):
                value_str = value_str + re_key_group[i]+","
                # 去除[]后用逗号分隔 [1,10,attackId],[1,40,attackId]变为"1, 10, 'attackId', 1, 40, 'attackId'"
                i+=2
        type_str_or_list = schema.get("type")
        if type(type_str_or_list) != list:
            # 简单数据
            return self.format_dispatch(type_str_or_list, value_str)
        else:
            if schema.get("special_key") == None:
                # 策划的配置太飘逸了
                value_new = value_str
                if value_str[-1:] == ";" or value_str[-1:]==",":
                    value_new = value_str[:-1]
                mark = ","
                if value_new.find(";") >= 0:
                    mark = ";"
                type_len = len(type_str_or_list)
                value_one_dim = [self.format_dispatch(type_str_or_list[i%type_len], value_new.split(mark)[i]) for i in range(len(value_new.split(mark)))]
                # 例：type:[int,int,string]
                #     value:[1,10,attackId],[1,40,attackId]
                #     每3个一循环，当前下标除以3余数的type
                return reshape_list(value_one_dim,type_len)
            else:
                return self.format_complex_data(schema.get("special_key"), type_str_or_list, value_str)

    # 考虑到可以单表更新，不把分开的表合并到一个对象中
    # 采用这种链式查找的方式获取对应的数据
    def get_record(self, tag):
        record = self._model.get(tag)
        if record == None:
            for model in self._chain:
                record = model.get(tag)
                if record:
                    return record
        else:
            return record

    def get_dict_record(self, tag):
        record = self.get_record(tag)
        if record:
            r_len = len(record)
            ret_dict = {}
            for index in range(len(self._key_list)):
                if index >= r_len:
                    ret_dict.update({self._key_list[index]:None})
                else:
                    ret_dict.update({self._key_list[index]:record[index]})
            return ret_dict
        else:
            return None

    def check_max_column(self):
        for sheet_name in self._valid_sheet:
            work_sheet = self._work_book.get_sheet_by_name(sheet_name)
            if work_sheet.max_column >= MAX_COLUMN:
                self.raise_exception("Max Col Error max column: {}".format(work_sheet.max_column))

    def check_link(self):
        if self._table_name.find("Translate") >= 0 or self._table_name.find("ConfigValue") >= 0:
            return 
        for sheet_name in self._valid_sheet:
            self._current_sheet = sheet_name
            work_sheet = self._work_book.get_sheet_by_name(sheet_name)
            valid_cols = self._get_valid_column(sheet_name)
            for col_idx in valid_cols:
                key = work_sheet.cell(row=KEY_ROW, column=col_idx).value
                self._current_key = key
                colunm_cells = self.query_cells_by_column(sheet_name, col_idx)
                value = ""
                for cell in colunm_cells:
                    if cell.value != None:
                        value = cell.value
                        break
                if self._need_link(key, value):
                    link = work_sheet.cell(row=LINK_ROW, column=col_idx).value
                    if link==None or link.strip() == "":
                        self.raise_exception("Link Error")

    def _need_link(self, key, value):
        black_list = ["tid#", "reward_"]
        if type(value) == str:
            for b in black_list:
                if value.find(b) >= 0 :
                    return True
        

    def query_cells_by_column(self, sheet_name, col_num):
        value_row = self._value_dict.get(sheet_name)
        limit = {
            "min_row":value_row,
            "max_col":col_num,
            "min_col":col_num
        }
        work_sheet = self._work_book.get_sheet_by_name(sheet_name)
        for col in work_sheet.iter_cols(**limit):
            return col


    def get_all_tags(self):
        return self._all_tags
    
    def raise_exception(self, msg):
        self._exception_info.append([self._table_name, self._current_sheet,
             self._current_tag, self._current_key, msg])
    
    def get_exception_info(self):
        return self._exception_info

    def chain(self, excel_obj):
        self._chain.append(excel_obj._model)

    def get_link_excel(self, key):
        link_table = self._schema.get("link_table")
        if link_table and link_table.get(key):
            excel_obj = Container.find_table(link_table.get(key))
            return excel_obj
        return None
    
    def has(self, tag):
        if self._model.get(tag) == None:
            for model in self._chain:
                if model.get(tag) != None:
                    return True
            return False
        else:
            return True
    
    def has_value_with_key(self, key, value):
        for sheet_name in self._valid_sheet:
            sheet_schema = self._schema[sheet_name]
            if sheet_schema == None:
                return False
            key_schema = sheet_schema.get(key)
            if key_schema == None:
                return False
            cells = self.query_cells_by_column(sheet_name, key_schema["column"])
            for cell in cells:
                cell_value = self.format_cell(key_schema, cell.value)
                if cell_value == value:
                    return True

        return True
    def get_value_by_key(self, tag, key):
        record = self.get_record(tag)
        if record:
            return record.get_value(key)
        return None
    def foreach(self, fn):
        for tag in self._all_tags:
            fn(self._model.get(tag))
    def get_records(self):
        return self._all_records

#TODO link表的处理
if __name__ == "__main__":
    # print("in test")
    start = time.time()
    excel_name = "Item.xlsx"
    te = DreamerExcel(excel_name)
    te.load_data()
